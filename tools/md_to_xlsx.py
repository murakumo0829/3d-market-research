import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def main(md_path, xlsx_path, date_str):
    with open(md_path, "r", encoding="utf-8") as f:
        text = f.read()

    # Split into level-2 sections by lines starting with "## "
    lines = text.split("\n")
    sections = []  # list of (title, body_lines)
    current_title = None
    current_body = []
    for line in lines:
        if line.startswith("## ") and not line.startswith("### "):
            if current_title is not None:
                sections.append((current_title, current_body))
            current_title = line[3:].strip()
            current_body = []
        elif current_title is not None:
            current_body.append(line)
    if current_title is not None:
        sections.append((current_title, current_body))

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    sub_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    src_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    link_re = re.compile(r"^-\s*\[(.+?)\]\((https?://[^\)]+)\)\s*$")

    for title, body in sections:
        # sheet name: strip leading number, limit length, remove invalid chars
        sheet_title = re.sub(r"^\d+\.\s*", "", title)
        sheet_title = re.sub(r"[\\/*\[\]:?]", "", sheet_title)[:31]
        ws = wb.create_sheet(title=sheet_title)
        ws.column_dimensions["A"].width = 110
        ws.column_dimensions["B"].width = 55

        row = 1
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = header_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 2

        in_source_block = False
        source_rows_start = None

        for raw in body:
            line = raw.strip()
            if not line:
                continue
            if line.startswith("### "):
                sub = line[4:].strip()
                if sub == "出典":
                    in_source_block = True
                    cell = ws.cell(row=row, column=1, value="出典")
                    cell.font = Font(bold=True, size=12)
                    cell.fill = src_fill
                    ws.cell(row=row, column=2, value="URL").font = Font(bold=True)
                    ws.cell(row=row, column=2).fill = src_fill
                    row += 1
                    continue
                else:
                    in_source_block = False
                    cell = ws.cell(row=row, column=1, value=sub)
                    cell.font = Font(bold=True, size=12)
                    cell.fill = sub_fill
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                    row += 1
                    continue

            if in_source_block:
                m = link_re.match(line)
                if m:
                    ws.cell(row=row, column=1, value=m.group(1))
                    url_cell = ws.cell(row=row, column=2, value=m.group(2))
                    url_cell.hyperlink = m.group(2)
                    url_cell.font = Font(color="0563C1", underline="single")
                    row += 1
                    continue
                else:
                    ws.cell(row=row, column=1, value=line)
                    row += 1
                    continue

            # regular content line
            c = ws.cell(row=row, column=1, value=line)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = min(15 * (1 + len(line) // 90), 200)
            row += 1

        ws.freeze_panes = "A3"

    wb.save(xlsx_path)
    print(f"Saved: {xlsx_path} ({len(sections)} sections)")

if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2], sys.argv[3])
